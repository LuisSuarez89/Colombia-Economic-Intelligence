"""Extract published GDP values using a validated workbook inspection."""
from __future__ import annotations

import logging
from collections.abc import Sequence
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal
from hashlib import sha256
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook

from .dane_pib_resolver import SourceManifest

LOGGER = logging.getLogger(__name__)


class ExtractionError(RuntimeError):
    """Base exception for extraction failures."""


class InvalidInspectionError(ExtractionError):
    """Raised when the supplied inspection cannot authorize extraction."""


class WorkbookMismatchError(ExtractionError):
    """Raised when the workbook does not match its source metadata."""


class ExtractionStructureError(ExtractionError):
    """Raised when an inspected table cannot be resolved safely."""


class UnexpectedCellValueError(ExtractionError):
    """Raised when a required value cell contains incompatible data."""


class DuplicateExtractionError(ExtractionError):
    """Raised when two inspected cells produce the same extraction key."""


@dataclass(frozen=True)
class ExtractionConfiguration:
    """Explicit policies that affect one extraction run."""

    extractor_version: str = "1.0"
    configuration_version: str = "1.0"
    hash_workbook: bool = False
    discard_empty_values: bool = True
    discard_unexpected_text: bool = True


@dataclass(frozen=True)
class ExtractionInput:
    """Workbook and the inspection that describes its readable regions."""

    workbook_path: Path
    inspection: InspectionResultLike
    source_manifest: SourceManifest | None = None
    configuration: ExtractionConfiguration = field(default_factory=ExtractionConfiguration)


@dataclass(frozen=True)
class ExtractedGDPRecord:
    source_file: str
    source_sha256: str | None
    source_sheet: str
    cuadro: int
    table_id: str
    table_region: str
    series_type: str | None
    aggregation_level: int | None
    indicator: str
    period: str
    period_original: str | None
    year: int
    quarter: int
    status: str | None
    classification_level: str | None
    activity_code: str | None
    activity_name: str
    entity_kind: str | None
    total_type: str | None
    value: Decimal
    unit: str | None
    price_basis: str | None
    source_row: int
    source_column: str
    raw_value: object


@dataclass(frozen=True)
class ExtractionMetadata:
    source_manifest: SourceManifest | None
    inspector_summary: dict[str, int]
    configuration_version: str
    extractor_version: str
    started_at: datetime
    completed_at: datetime
    tables_processed: int
    records_extracted: int
    records_discarded: int
    null_values: int
    duplicates: int
    errors: tuple[str, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class ExtractionResult:
    records: Sequence[ExtractedGDPRecord]
    metadata: ExtractionMetadata


class _TableLike(Protocol):
    table_id: str
    indicator: str | None
    periods: Sequence[object]
    activities: Sequence[object]


class InspectionResultLike(Protocol):
    is_valid: bool
    sheets: Sequence[object]
    tables: Sequence[_TableLike]


def _text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _source_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _period_original(period: Any) -> str | None:
    year_header = getattr(period, "year_header", None)
    quarter_header = getattr(period, "quarter_header", None)
    if year_header is None and quarter_header is None:
        return None
    return "-".join(part for part in (_text(year_header), _text(quarter_header)) if part)


def _decimal_value(value: object) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    return None


def _table_region(table: _TableLike) -> str:
    return _text(getattr(table, "region_id", None) or getattr(table, "table_region", None) or table.table_id)


def _tables_for_sheet(inspection: Any, sheet: Any) -> list[_TableLike]:
    tables = list(getattr(sheet, "detected_tables", ()))
    if tables:
        return tables
    sheet_name = getattr(sheet, "name", None)
    return [table for table in getattr(inspection, "tables", ()) if getattr(table, "sheet_name", None) == sheet_name]


class PIBExtractor:
    """Read only the value cells authorized by ``InspectionResult``."""

    @classmethod
    def extract(cls, extraction_input: ExtractionInput) -> ExtractionResult:
        started_at = datetime.now(timezone.utc)
        configuration = extraction_input.configuration
        path = Path(extraction_input.workbook_path)
        inspection = extraction_input.inspection
        if inspection is None or not getattr(inspection, "is_valid", False):
            raise InvalidInspectionError("A valid InspectionResult is required")
        if not path.is_file():
            raise ExtractionError(f"Workbook does not exist: {path}")
        source_hash = extraction_input.source_manifest.sha256 if extraction_input.source_manifest else None
        actual_hash = _source_sha256(path) if (source_hash or configuration.hash_workbook) else None
        if source_hash and actual_hash != source_hash:
            raise WorkbookMismatchError(f"Workbook hash does not match manifest: {path}")
        if extraction_input.source_manifest and extraction_input.source_manifest.filename != path.name:
            raise WorkbookMismatchError(f"Workbook filename does not match manifest: {path.name}")

        workbook = None
        records: list[ExtractedGDPRecord] = []
        warnings: list[str] = []
        seen: set[tuple[object, ...]] = set()
        discarded = 0
        null_values = 0
        try:
            workbook = load_workbook(path, data_only=False, read_only=False)
            sheets = [sheet for sheet in getattr(inspection, "sheets", ()) if getattr(sheet, "sheet_type", None) == "CUADRO"]
            if not sheets:
                raise ExtractionStructureError("InspectionResult contains no CUADRO sheets")
            tables_processed = 0
            for sheet_info in sheets:
                sheet_name = _text(getattr(sheet_info, "name", None))
                if not sheet_name:
                    raise ExtractionStructureError("Inspected sheet has no name")
                if sheet_name not in workbook.sheetnames:
                    raise ExtractionStructureError(f"Inspected sheet is absent: {sheet_name}")
                worksheet = workbook[sheet_name]
                tables = _tables_for_sheet(inspection, sheet_info)
                if not tables:
                    raise ExtractionStructureError(f"Inspected sheet has no tables: {sheet_name}")
                for table in tables:
                    tables_processed += 1
                    indicator = getattr(table, "indicator", None)
                    if not indicator:
                        raise ExtractionStructureError(f"Table has no resolved indicator: {sheet_name}")
                    periods = list(getattr(table, "periods", ()))
                    activities = list(getattr(table, "activities", ()))
                    if not periods or not activities:
                        raise ExtractionStructureError(f"Table has no resolved periods or activities: {table.table_id}")
                    for activity in activities:
                        row_number = getattr(activity, "row_number", None)
                        activity_name = _text(getattr(activity, "concept", None) or getattr(activity, "activity_name", None))
                        if not isinstance(row_number, int) or not activity_name:
                            raise ExtractionStructureError(f"Invalid activity in {table.table_id}")
                        for period in periods:
                            column = getattr(period, "column", None)
                            if not isinstance(column, str) or not column:
                                raise ExtractionStructureError(f"Invalid period column in {table.table_id}")
                            cell = worksheet[f"{column}{row_number}"]
                            value = _decimal_value(cell.value)
                            location = f"{sheet_name}!{column}{row_number}"
                            if value is None:
                                is_empty = cell.value is None or _text(cell.value) == ""
                                if is_empty:
                                    null_values += 1
                                if is_empty and configuration.discard_empty_values:
                                    discarded += 1
                                    warnings.append(f"Discarded empty value at {location}")
                                    continue
                                if configuration.discard_unexpected_text:
                                    discarded += 1
                                    warnings.append(f"Discarded non-numeric value at {location}")
                                    continue
                                raise UnexpectedCellValueError(f"Unexpected value at {location}: {cell.value!r}")
                            period_label = _text(getattr(period, "label", None))
                            key = (source_hash or path.name, sheet_name, table.table_id, indicator, period_label, row_number, _text(getattr(activity, "classification_code", None) or activity_name))
                            if key in seen:
                                raise DuplicateExtractionError(f"Duplicate extraction key at {location}")
                            seen.add(key)
                            record = ExtractedGDPRecord(
                                source_file=path.name,
                                source_sha256=source_hash or actual_hash,
                                source_sheet=sheet_name,
                                cuadro=int(getattr(sheet_info, "cuadro_number", 0) or 0),
                                table_id=table.table_id,
                                table_region=_table_region(table),
                                series_type=getattr(sheet_info, "series", None),
                                aggregation_level=getattr(sheet_info, "aggregation_level", None),
                                indicator=indicator,
                                period=period_label,
                                period_original=_period_original(period),
                                year=int(period.year),
                                quarter=int(period.quarter),
                                status=getattr(period, "status", None),
                                classification_level=getattr(activity, "classification_level", None),
                                activity_code=getattr(activity, "classification_code", None),
                                activity_name=activity_name,
                                entity_kind="total" if getattr(activity, "is_total", False) else "activity",
                                total_type=getattr(activity, "total_type", None),
                                value=value,
                                unit=getattr(table, "unit", None),
                                price_basis=getattr(table, "price_basis", None),
                                source_row=row_number,
                                source_column=column,
                                raw_value=cell.value,
                            )
                            records.append(record)
        finally:
            if workbook is not None:
                workbook.close()
        completed_at = datetime.now(timezone.utc)
        records.sort(key=lambda record: (record.cuadro, record.table_id, record.source_row, record.source_column))
        metadata = ExtractionMetadata(
            source_manifest=extraction_input.source_manifest,
            inspector_summary={"sheets": len(getattr(inspection, "sheets", ())), "tables": tables_processed, "periods": sum(len(getattr(table, "periods", ())) for table in getattr(inspection, "tables", ()))},
            configuration_version=configuration.configuration_version,
            extractor_version=configuration.extractor_version,
            started_at=started_at,
            completed_at=completed_at,
            tables_processed=tables_processed,
            records_extracted=len(records),
            records_discarded=discarded,
            null_values=null_values,
            duplicates=0,
            errors=(),
            warnings=tuple(warnings),
        )
        LOGGER.info("Extracted %d records from %d tables in %s", len(records), tables_processed, path.name)
        return ExtractionResult(records=tuple(records), metadata=metadata)