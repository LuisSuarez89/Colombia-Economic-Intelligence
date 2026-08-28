"""Structural inspection of DANE quarterly GDP production workbooks."""
from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from itertools import pairwise
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXPECTED_SHEET_COUNT = 7
EXPECTED_AGGREGATIONS = (12, 25, 61)
QUARTERS = {"I": 1, "II": 2, "III": 3, "IV": 4}
TOTAL_TYPES = {
    "valor agregado bruto": "VAB",
    "impuestos menos subvenciones sobre los productos": "IMPUESTOS_MENOS_SUBVENCIONES",
    "producto interno bruto": "PIB",
}
INDICATOR_TITLES = {
    "CRECIMIENTO_ANUAL": ("tasa de crecimiento anual",),
    "CRECIMIENTO_TRIMESTRAL": ("tasa de crecimiento trimestral",),
    "CRECIMIENTO_ANO_CORRIDO": ("tasa de crecimiento año corrido", "tasa de crecimiento ano corrido"),
}
EXPECTED_INDICATORS = {
    "ORIGINAL": ("NIVEL", "CRECIMIENTO_ANUAL", "CRECIMIENTO_ANO_CORRIDO"),
    "AJUSTADA": ("NIVEL", "CRECIMIENTO_TRIMESTRAL", "CRECIMIENTO_ANO_CORRIDO"),
}
EXPECTED_START = {
    "NIVEL": (2005, 1),
    "CRECIMIENTO_ANUAL": (2006, 1),
    "CRECIMIENTO_ANO_CORRIDO": (2006, 1),
    "CRECIMIENTO_TRIMESTRAL": (2005, 2),
}


@dataclass(frozen=True)
class ValidationResult:
    code: str
    severity: str
    message: str
    expected: Any = None
    actual: Any = None


@dataclass(frozen=True)
class Period:
    year: int
    quarter: int
    status: str | None
    column: str
    label: str
    year_header: str | None = None
    quarter_header: str | None = None


@dataclass(frozen=True)
class ActivityRow:
    row_number: int
    classification_level: str
    classification_code: str | None
    concept: str
    is_total: bool
    total_type: str | None


@dataclass
class TableInspection:
    table_id: str
    indicator: str | None
    title: str
    header_row: int | None
    period_row: int | None
    data_start_row: int | None
    data_end_row: int | None
    first_period_column: str | None
    last_period_column: str | None
    period_count: int
    activity_row_count: int
    validation_status: str = "VALID"
    periods: list[Period] = field(default_factory=list)
    activities: list[ActivityRow] = field(default_factory=list)
    unit: str | None = None
    price_basis: str | None = None


@dataclass(frozen=True)
class TableRegion:
    title_row: int
    title: str
    indicator: str | None
    header_row: int
    period_row: int
    periods: list[Period]
    invalid_status: list[int]
    invalid_quarters: list[int]
    first_period_column: int


@dataclass
class SheetInspection:
    name: str
    sheet_type: str
    cuadro_number: int | None
    series: str | None
    aggregation_level: int | None
    expected_tables: int
    detected_tables: list[TableInspection]
    header_structure: dict[str, Any] | None
    validation_status: str = "VALID"


@dataclass(frozen=True)
class WorkbookMetadata:
    filename: str | None
    sheet_count: int | None
    expected_sheet_count: int
    source: str | None
    publication_date: str | None
    reference_year: int | None
    detected_series: tuple[str, ...] | None


@dataclass
class InspectionResult:
    workbook: WorkbookMetadata
    sheets: list[SheetInspection]
    tables: list[TableInspection]
    periods: list[Period]
    activities: list[ActivityRow]
    validations: list[ValidationResult]
    warnings: list[ValidationResult]

    @property
    def is_valid(self) -> bool:
        return not any(item.severity == "ERROR" for item in self.validations)


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value)).strip() if value is not None else ""


def _normalized(value: Any) -> str:
    return _text(value).casefold()


def _find_aggregation(texts: Iterable[str]) -> int | None:
    joined = " ".join(texts).casefold()
    for level in reversed(EXPECTED_AGGREGATIONS):
        if re.search(rf"\b{level}\s+agrupaciones\b", joined):
            return level
    return None


def _find_series(texts: Iterable[str]) -> str | None:
    joined = " ".join(texts).casefold()
    if "ajustad" in joined and "efecto estacional" in joined and "calendario" in joined:
        return "AJUSTADA"
    if "datos originales" in joined or "originales" in joined:
        return "ORIGINAL"
    return None


def _parse_year(value: Any) -> tuple[int, str | None] | None:
    match = re.fullmatch(r"(20\d{2})(pr|p)?", _text(value).casefold())
    return (int(match.group(1)), match.group(2)) if match else None


def _merged_cache(sheet: Any) -> dict[tuple[int, int], Any]:
    cache: dict[tuple[int, int], Any] = {}
    for merged in sheet.merged_cells.ranges:
        value = sheet.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                cache[(row, column)] = value
    return cache


def _period_header(sheet: Any, header_row: int, merged: dict[tuple[int, int], Any]) -> tuple[int, list[Period], list[int], list[int]] | None:
    period_row = header_row + 1
    periods: list[Period] = []
    invalid_status: list[int] = []
    invalid_quarters: list[int] = []
    for column in range(1, sheet.max_column + 1):
        year_value = merged.get((header_row, column), sheet.cell(header_row, column).value)
        quarter_header = _text(sheet.cell(period_row, column).value).upper()
        next_year = merged.get((header_row, column + 1), sheet.cell(header_row, column + 1).value) if column < sheet.max_column else None
        next_quarter = _text(sheet.cell(period_row, column + 1).value).upper() if column < sheet.max_column else ""
        if not _text(year_value) and quarter_header == "I" and next_quarter == "II" and _parse_year(next_year):
            year_value = next_year
        parsed_year = _parse_year(year_value)
        if parsed_year and quarter_header in QUARTERS:
            year, status = parsed_year
            periods.append(Period(year, QUARTERS[quarter_header], status, get_column_letter(column), f"{year}-{quarter_header}" + (f"-{status}" if status else ""), _text(year_value), quarter_header))
        elif re.fullmatch(r"20\d{2}[A-Za-z]+", _text(year_value)) and quarter_header in QUARTERS:
            invalid_status.append(column)
        elif re.fullmatch(r"20\d{2}(?:p|pr)?", _text(year_value), re.IGNORECASE) and quarter_header and quarter_header not in QUARTERS:
            invalid_quarters.append(column)
    return (period_row, periods, invalid_status, invalid_quarters) if periods or invalid_status or invalid_quarters else None


def _indicator(texts: Iterable[str]) -> str | None:
    joined = " ".join(_normalized(text) for text in texts)
    return next((indicator for indicator, titles in INDICATOR_TITLES.items() if any(title in joined for title in titles)), None)


def _table_regions(sheet: Any, merged: dict[tuple[int, int], Any]) -> list[TableRegion]:
    header_rows = [row for row in range(1, sheet.max_row) if _period_header(sheet, row, merged)]
    regions: list[TableRegion] = []
    for index, header_row in enumerate(header_rows):
        header_info = _period_header(sheet, header_row, merged)
        if not header_info:
            continue
        period_row, periods, invalid_status, invalid_quarters = header_info
        first_period = next((period for period in periods), None)
        if first_period is None:
            continue
        first_column = next(column for column in range(1, sheet.max_column + 1) if get_column_letter(column) == first_period.column)
        next_header = header_rows[index + 1] if index + 1 < len(header_rows) else sheet.max_row + 1
        data_rows = range(period_row + 1, next_header)
        first_data_column = next(
            (column for column in range(first_column, sheet.max_column + 1)
             if any(isinstance(sheet.cell(row, column).value, (int, float)) for row in data_rows)),
            first_column,
        )
        context_rows: list[tuple[int, list[str]]] = []
        for row in range(max(1, header_row - 6), header_row):
            values = [_text(sheet.cell(row, column).value) for column in range(1, first_column)]
            values = [value for value in values if value]
            if values:
                context_rows.append((row, values))
        indicator = next((_indicator(values) for _, values in reversed(context_rows) if _indicator(values)), None)
        if indicator is None and index == 0 and any("miles de millones de pesos" in value.casefold() for _, values in context_rows for value in values):
            indicator = "NIVEL"
        title_row, title = (context_rows[-1] if context_rows else (header_row, ""))
        periods = [
            period for period in periods
            if next(column for column in range(1, sheet.max_column + 1) if get_column_letter(column) == period.column) >= first_data_column
        ]
        regions.append(TableRegion(title_row, " ".join(title), indicator, header_row, period_row, periods, invalid_status, invalid_quarters, first_column))
    return regions


def _activities(sheet: Any, start: int, end: int, first_period_column: int, aggregation: int) -> list[ActivityRow]:
    result: list[ActivityRow] = []
    for row in range(start, end + 1):
        values = [_text(sheet.cell(row, column).value) for column in range(1, first_period_column)]
        values = [value for value in values if value]
        if not values:
            continue
        concept = values[-1]
        if concept.casefold().startswith(("fuente:", "actualizado", "pprovisional", "prpreliminar")):
            continue
        code = values[-2] if len(values) > 1 else None
        total_type = next((kind for name, kind in TOTAL_TYPES.items() if name in concept.casefold()), None)
        result.append(ActivityRow(row, f"CIIU_{aggregation}", code, concept, total_type is not None, total_type))
    return result


class PIBWorkbookInspector:
    @classmethod
    def inspect(cls, path: str | Path) -> InspectionResult:
        workbook_path = Path(path)
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        validations: list[ValidationResult] = []
        warnings: list[ValidationResult] = []
        sheets: list[SheetInspection] = []
        for sheet in workbook.worksheets:
            all_text = [_text(value) for row in sheet.iter_rows(values_only=True) for value in row if _text(value)]
            normalized_name = _normalized(sheet.title)
            if normalized_name in {"índice", "indice"}:
                sheets.append(SheetInspection(sheet.title, "INDEX", None, None, None, 0, [], None))
                continue
            series = _find_series(all_text)
            aggregation = _find_aggregation(all_text)
            match = re.search(r"\bcuadro\s+([1-6])\b", normalized_name)
            cuadro_number = int(match.group(1)) if match else None
            expected_series = "ORIGINAL" if cuadro_number in (1, 2, 3) else "AJUSTADA" if cuadro_number in (4, 5, 6) else None
            expected_aggregation = {1: 12, 2: 25, 3: 61}.get(((cuadro_number - 1) % 3 + 1) if cuadro_number else 0)
            if cuadro_number and (series != expected_series or aggregation != expected_aggregation):
                validations.append(ValidationResult("UNEXPECTED_STRUCTURE", "ERROR", f"{sheet.title} contradice su combinación declarada"))
            tables: list[TableInspection] = []
            for index, region in enumerate(_table_regions(sheet, _merged_cache(sheet))[:3], start=1):
                expected_indicator = EXPECTED_INDICATORS.get(series or "", ())[index - 1] if index <= len(EXPECTED_INDICATORS.get(series or "", ())) else None
                resolved_indicator = expected_indicator if index == 1 and region.indicator is None else region.indicator
                next_title = _table_regions(sheet, _merged_cache(sheet))[index].title_row if index < len(_table_regions(sheet, _merged_cache(sheet))) else sheet.max_row + 1
                activities = _activities(sheet, region.period_row + 1, next_title - 1, region.first_period_column, aggregation or 0)
                tables.append(TableInspection(f"Cuadro {cuadro_number or sheet.title}-T{index}", resolved_indicator, region.title, region.header_row, region.period_row, activities[0].row_number if activities else None, activities[-1].row_number if activities else None, region.periods[0].column if region.periods else None, region.periods[-1].column if region.periods else None, len(region.periods), len(activities), "VALID", region.periods, activities))
                if resolved_indicator != expected_indicator:
                    validations.append(ValidationResult("UNEXPECTED_INDICATOR", "ERROR", f"Indicador incompatible en {sheet.title}"))
                if region.invalid_status:
                    validations.append(ValidationResult("UNKNOWN_STATUS", "ERROR", f"Estado temporal desconocido en {sheet.title}"))
                if region.invalid_quarters:
                    validations.append(ValidationResult("INVALID_PERIOD", "ERROR", f"Trimestre inválido en {sheet.title}"))
                expected_start = EXPECTED_START.get(resolved_indicator or "")
                if expected_start and region.periods and (region.periods[0].year, region.periods[0].quarter) != expected_start:
                    validations.append(ValidationResult("INVALID_PERIOD", "ERROR", f"Horizonte inicial inválido en {sheet.title}"))
                for previous, current in pairwise(region.periods):
                    expected_next = (previous.year + 1, 1) if previous.quarter == 4 else (previous.year, previous.quarter + 1)
                    if (current.year, current.quarter) != expected_next:
                        validations.append(ValidationResult("INVALID_PERIOD", "ERROR", f"Períodos no secuenciales en {sheet.title}"))
            if len(tables) != 3:
                validations.append(ValidationResult("EXPECTED_TABLE_COUNT", "ERROR", f"{sheet.title} debe tener tres tablas"))
            sheets.append(SheetInspection(sheet.title, "CUADRO", cuadro_number, series, aggregation, 3, tables, {"merged_ranges": tuple(str(item) for item in sheet.merged_cells.ranges)}))
        metadata_text = "\n".join(_text(value) for sheet in workbook.worksheets for row in sheet.iter_rows(values_only=True) for value in row if _text(value))
        reference = re.search(r"año de referencia\s+(20\d{2})", metadata_text, re.IGNORECASE)
        date = re.search(r"actualizado el\s+([^\n]+)", metadata_text, re.IGNORECASE)
        source = re.search(r"fuente:\s*([^\n]+)", metadata_text, re.IGNORECASE)
        result = InspectionResult(WorkbookMetadata(workbook_path.name, len(workbook.sheetnames), EXPECTED_SHEET_COUNT, source.group(1).strip() if source else None, date.group(1).strip() if date else None, int(reference.group(1)) if reference else None, tuple(sorted({sheet.series for sheet in sheets if sheet.series})) or None), sheets, [table for sheet in sheets for table in sheet.detected_tables], [period for sheet in sheets for table in sheet.detected_tables for period in table.periods], [activity for sheet in sheets for table in sheet.detected_tables for activity in table.activities], validations, warnings)
        if len(workbook.sheetnames) != EXPECTED_SHEET_COUNT:
            result.validations.append(ValidationResult("EXPECTED_SHEET_COUNT", "ERROR", "Cantidad de hojas inesperada"))
        return result