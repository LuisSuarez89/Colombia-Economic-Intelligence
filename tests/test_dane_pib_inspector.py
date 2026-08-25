from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from colombia_economic_intelligence.sources.dane_pib_inspector import (
    PIBWorkbookInspector,
)


def _workbook(path: Path, future: bool = False) -> None:
    workbook = Workbook()
    index = workbook.active
    index.title = "Índice"
    index["A1"] = "Series encadenadas de volumen con año de referencia 2015"
    index["A2"] = "Fuente: DANE, PIB_T"
    index["A3"] = "Actualizado el 18 de agosto de 2026"
    combinations = [("Datos originales", 12), ("Datos originales", 25), ("Datos originales", 61), ("Datos ajustados por efecto estacional y calendario", 12), ("Datos ajustados por efecto estacional y calendario", 25), ("Datos ajustados por efecto estacional y calendario", 61)]
    for number, (series, aggregation) in enumerate(combinations, 1):
        sheet = workbook.create_sheet(f"Cuadro {number}")
        sheet["A1"] = series
        sheet["A2"] = f"Secciones CIIU Rev. 4 A.C. / {aggregation} agrupaciones"
        growth = "Tasa de crecimiento trimestral" if number >= 4 else "Tasa de crecimiento anual"
        starts = [(2005, 1), (2005, 2) if number >= 4 else (2006, 1), (2006, 1)]
        titles = ((3, "Miles de millones de pesos"), (12, growth), (20, "Tasa de crecimiento año corrido"))
        for table_index, (row, title) in enumerate(titles):
            sheet.cell(row, 1).value = title
            sheet.cell(row + 1, 1).value = "Sección"
            year, quarter = starts[table_index]
            last_year, last_quarter = (2027, 1) if future else (2026, 2)
            column = 3
            while (year, quarter) <= (last_year, last_quarter):
                first_column = column
                year_status = "pr" if year >= 2026 else "p" if year == 2025 else None
                while quarter <= 4 and (year, quarter) <= (last_year, last_quarter):
                    sheet.cell(row + 2, column).value = ("I", "II", "III", "IV")[quarter - 1]
                    column += 1
                    quarter += 1
                sheet.cell(row + 1, first_column).value = f"{year}{year_status or ''}"
                if column - first_column > 1:
                    sheet.merge_cells(start_row=row + 1, start_column=first_column, end_row=row + 1, end_column=column - 1)
                year += 1
                quarter = 1
            sheet.cell(row + 3, 1).value = "A"
            sheet.cell(row + 3, 2).value = "Agricultura"
        sheet["B7"] = "Valor agregado bruto"
        sheet["B8"] = "Impuestos menos subvenciones sobre los productos"
        sheet["B9"] = "Producto Interno Bruto"
    workbook.save(path)


def test_inspector_expands_merged_years_and_detects_structure(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path)
    result = PIBWorkbookInspector.inspect(path)
    assert result.workbook.sheet_count == 7
    assert result.workbook.reference_year == 2015
    assert len(result.tables) == 18
    assert result.is_valid
    assert result.periods[0].label == "2005-I"
    assert result.periods[1].label == "2005-II"
    assert result.workbook.source == "DANE, PIB_T"
    assert result.workbook.publication_date == "18 de agosto de 2026"
    assert result.sheets[1].series == "ORIGINAL"
    assert result.sheets[4].series == "AJUSTADA"
    assert result.sheets[1].aggregation_level == 12
    assert result.activities[1].total_type == "VAB"
    assert [table.indicator for table in result.sheets[1].detected_tables] == ["NIVEL", "CRECIMIENTO_ANUAL", "CRECIMIENTO_ANO_CORRIDO"]
    assert [table.indicator for table in result.sheets[4].detected_tables] == ["NIVEL", "CRECIMIENTO_TRIMESTRAL", "CRECIMIENTO_ANO_CORRIDO"]
    assert {period.status for period in result.periods if period.year == 2025} == {"p"}
    assert {period.status for period in result.periods if period.year == 2026} == {"pr"}


def test_future_last_period_is_allowed(tmp_path: Path) -> None:
    path = tmp_path / "future.xlsx"
    _workbook(path, future=True)
    result = PIBWorkbookInspector.inspect(path)
    assert result.is_valid
    assert result.periods[-1].label == "2027-I-pr"


def test_indicator_detection_is_local_to_table_region(tmp_path: Path) -> None:
    path = tmp_path / "horizontal-regions.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    sheet = workbook["Cuadro 1"]
    sheet["I1"] = "Índice"
    sheet["I2"] = "Tasa de crecimiento anual"
    sheet["I3"] = "Tasa de crecimiento año corrido"
    workbook.save(path)

    result = PIBWorkbookInspector.inspect(path)

    assert result.is_valid
    assert [table.indicator for table in result.sheets[1].detected_tables] == [
        "NIVEL",
        "CRECIMIENTO_ANUAL",
        "CRECIMIENTO_ANO_CORRIDO",
    ]


def test_real_dane_workbook_integration() -> None:
    path = Path("local_data/anex-ProduccionConstantes-IItrim2026.xlsx")
    if not path.exists():
        pytest.skip("real DANE workbook is not available locally")

    result = PIBWorkbookInspector.inspect(path)

    assert result.is_valid
    assert result.workbook.sheet_count == 7
    assert sum(sheet.sheet_type == "CUADRO" for sheet in result.sheets) == 6
    assert len(result.tables) == 18
    expected = [
        ["NIVEL", "CRECIMIENTO_ANUAL", "CRECIMIENTO_ANO_CORRIDO"],
        ["NIVEL", "CRECIMIENTO_ANUAL", "CRECIMIENTO_ANO_CORRIDO"],
        ["NIVEL", "CRECIMIENTO_ANUAL", "CRECIMIENTO_ANO_CORRIDO"],
        ["NIVEL", "CRECIMIENTO_TRIMESTRAL", "CRECIMIENTO_ANO_CORRIDO"],
        ["NIVEL", "CRECIMIENTO_TRIMESTRAL", "CRECIMIENTO_ANO_CORRIDO"],
        ["NIVEL", "CRECIMIENTO_TRIMESTRAL", "CRECIMIENTO_ANO_CORRIDO"],
    ]
    assert [
        [table.indicator for table in sheet.detected_tables]
        for sheet in result.sheets
        if sheet.sheet_type == "CUADRO"
    ] == expected
    assert result.tables[0].periods[0].label == "2005-I"
    assert result.tables[1].periods[0].label == "2006-I"
    assert result.tables[10].periods[0].label == "2005-II"
    assert result.tables[14].periods[-1].label == "2026-II-pr"


def test_missing_metadata_returns_warning(tmp_path: Path) -> None:
    path = tmp_path / "metadata.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    workbook["Índice"]["A1"] = None
    workbook["Índice"]["A2"] = None
    workbook["Índice"]["A3"] = None
    workbook.save(path)
    result = PIBWorkbookInspector.inspect(path)
    assert result.workbook.reference_year is None
    assert result.workbook.source is None
    assert any(item.severity == "WARNING" for item in result.warnings)


def test_missing_table_is_error(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    sheet = workbook["Cuadro 1"]
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= 21 and merged.min_row <= 22:
            sheet.unmerge_cells(str(merged))
    for row in range(20, 24):
        for column in range(1, sheet.max_column + 1):
            sheet.cell(row, column).value = None
    workbook.save(path)
    result = PIBWorkbookInspector.inspect(path)
    assert any(item.code == "MISSING_TABLE" and item.severity == "ERROR" for item in result.validations)
    assert not result.is_valid


def test_unknown_indicator_is_error(tmp_path: Path) -> None:
    path = tmp_path / "unknown.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    workbook["Cuadro 1"]["A12"] = "Indicador desconocido"
    workbook.save(path)
    result = PIBWorkbookInspector.inspect(path)
    assert any(item.code == "UNEXPECTED_INDICATOR" and item.severity == "ERROR" for item in result.validations)


def test_invalid_status_and_quarter_are_errors(tmp_path: Path) -> None:
    path = tmp_path / "invalid.xlsx"
    _workbook(path)
    workbook = load_workbook(path)
    sheet = workbook["Cuadro 1"]
    for cell in sheet[14]:
        if cell.value == "IV":
            cell.value = "V"
    for merged in sheet.merged_cells.ranges:
        if merged.min_row == 4 and sheet.cell(merged.min_row, merged.min_col).value == "2026pr":
            sheet.cell(merged.min_row, merged.min_col).value = "2026x"
    workbook.save(path)
    result = PIBWorkbookInspector.inspect(path)
    codes = {item.code for item in result.validations}
    assert "UNKNOWN_STATUS" in codes
    assert "INVALID_PERIOD" in codes
    assert not result.is_valid