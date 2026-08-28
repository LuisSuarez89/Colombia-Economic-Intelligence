from decimal import Decimal
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import Workbook

from colombia_economic_intelligence.sources.dane_pib_extractor import (
    DuplicateExtractionError,
    ExtractionConfiguration,
    ExtractionInput,
    ExtractionStructureError,
    InvalidInspectionError,
    PIBExtractor,
    UnexpectedCellValueError,
)
from colombia_economic_intelligence.sources.dane_pib_inspector import (
    ActivityRow,
    InspectionResult,
    Period,
    PIBWorkbookInspector,
    SheetInspection,
    TableInspection,
    ValidationResult,
    WorkbookMetadata,
)


def _inspection(
    *,
    period: Period | None = None,
    activity: ActivityRow | None = None,
    table: TableInspection | None = None,
    sheet: SheetInspection | None = None,
) -> InspectionResult:
    period = period or Period(2026, 2, "pr", "B", "2026-II-pr", "2026pr", "II")
    activity = activity or ActivityRow(5, "CIIU_12", "045 - 047", "Agricultura", False, None)
    table = table or TableInspection(
        "Cuadro 1-T1", "NIVEL", "Miles de millones de pesos", 3, 4, 5, 5,
        "B", "B", 1, 1, "VALID", [period], [activity],
        "Miles de millones de pesos", "precios constantes"
    )
    sheet = sheet or SheetInspection("Cuadro 1", "CUADRO", 1, "ORIGINAL", 12, 1, [table], None)
    return InspectionResult(
        WorkbookMetadata("pib.xlsx", 1, 1, "DANE", None, 2015, ("ORIGINAL",)),
        [sheet], [table], [period], [activity], [], []
    )


def _workbook(path: Path, value: object = 123.45) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Cuadro 1"
    sheet["B5"] = cast(Any, value)
    workbook.save(path)


def _inspected_workbook(path: Path) -> None:
    workbook = Workbook()
    index = workbook.active
    assert index is not None
    index.title = "Índice"
    index["A1"] = "Series encadenadas de volumen con año de referencia 2015"
    index["A2"] = "Fuente: DANE, PIB_T"
    for number in range(1, 7):
        sheet = workbook.create_sheet(f"Cuadro {number}")
        sheet["A1"] = "Datos originales" if number <= 3 else "Datos ajustados por efecto estacional y calendario"
        sheet["A2"] = f"Secciones CIIU Rev. 4 A.C. / {12 if number in (1, 4) else 25 if number in (2, 5) else 61} agrupaciones"
        titles = ("Miles de millones de pesos", "Tasa de crecimiento trimestral" if number >= 4 else "Tasa de crecimiento anual", "Tasa de crecimiento año corrido")
        starts = ((2005, 1), (2005, 2) if number >= 4 else (2006, 1), (2006, 1))
        for offset, (title, start) in enumerate(zip(titles, starts)):
            title_row = (3, 12, 20)[offset]
            sheet.cell(title_row, 1).value = title
            sheet.cell(title_row + 1, 1).value = "Sección"
            year, quarter = start
            column = 3
            while (year, quarter) <= (2026, 2):
                first_column = column
                while quarter <= 4 and (year, quarter) <= (2026, 2):
                    sheet.cell(title_row + 2, column).value = ("I", "II", "III", "IV")[quarter - 1]
                    sheet.cell(title_row + 3, column).value = 10
                    column += 1
                    quarter += 1
                sheet.cell(title_row + 1, first_column).value = f"{year}{'pr' if year == 2026 else ''}"
                if column - first_column > 1:
                    sheet.merge_cells(start_row=title_row + 1, start_column=first_column, end_row=title_row + 1, end_column=column - 1)
                year += 1
                quarter = 1
            sheet.cell(title_row + 3, 1).value = "A"
            sheet.cell(title_row + 3, 2).value = "Agricultura"
    workbook.save(path)


def _input(path: Path, inspection: InspectionResult | None = None) -> ExtractionInput:
    return ExtractionInput(path, cast(Any, inspection or _inspection()))


def test_extracts_value_and_preserves_lineage(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path)

    result = PIBExtractor.extract(_input(path))

    record = result.records[0]
    assert record.value == Decimal("123.45")
    assert record.source_sheet == "Cuadro 1"
    assert (record.source_row, record.source_column) == (5, "B")
    assert record.activity_code == "045 - 047"
    assert record.unit == "Miles de millones de pesos"
    assert record.period_original == "2026pr-II"
    assert result.metadata.records_extracted == 1


@pytest.mark.parametrize(
    ("series", "indicator", "aggregation"),
    [
        ("ORIGINAL", "NIVEL", 12),
        ("ORIGINAL", "CRECIMIENTO_ANUAL", 25),
        ("ORIGINAL", "CRECIMIENTO_ANO_CORRIDO", 61),
        ("AJUSTADA", "CRECIMIENTO_TRIMESTRAL", 12),
    ],
)
def test_consumes_series_indicator_and_aggregation_from_inspection(
    tmp_path: Path, series: str, indicator: str, aggregation: int
) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path)
    inspection = _inspection()
    sheet = inspection.sheets[0]
    sheet.series = series
    sheet.aggregation_level = aggregation
    sheet.detected_tables[0].indicator = indicator

    record = PIBExtractor.extract(_input(path, inspection)).records[0]

    assert (record.series_type, record.indicator, record.aggregation_level) == (series, indicator, aggregation)


def test_preserves_zero_and_discards_empty_value(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path, 0)
    result = PIBExtractor.extract(_input(path))
    assert result.records[0].value == 0

    _workbook(path, None)
    result = PIBExtractor.extract(_input(path))
    assert result.records == ()
    assert result.metadata.null_values == 1
    assert result.metadata.records_discarded == 1


def test_preserves_p_status_and_future_period_without_hardcoding(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path)
    inspection = _inspection()
    inspection.sheets[0].series = "AJUSTADA"
    inspection.sheets[0].detected_tables[0].indicator = "CRECIMIENTO_ANO_CORRIDO"
    inspection.sheets[0].detected_tables[0].periods[0] = Period(year=2031, quarter=4, status="p", column="B", label="2031-IV-p", year_header="2031p", quarter_header="IV")

    record = PIBExtractor.extract(_input(path, inspection)).records[0]

    assert (record.period, record.year, record.quarter, record.status) == ("2031-IV-p", 2031, 4, "p")


def test_rejects_invalid_inspection_and_missing_table_context(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path)
    invalid = _inspection()
    invalid.validations.append(ValidationResult("INVALID", "ERROR", "invalid fixture"))
    with pytest.raises(InvalidInspectionError):
        PIBExtractor.extract(_input(path, invalid))

    empty = _inspection()
    empty.sheets[0].detected_tables = []
    empty.tables = []
    with pytest.raises(ExtractionStructureError):
        PIBExtractor.extract(_input(path, empty))


def test_rejects_unexpected_text_and_duplicate_keys(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path, "N.D.")
    strict = ExtractionInput(
        path,
        cast(Any, _inspection()),
        configuration=ExtractionConfiguration(discard_unexpected_text=False),
    )
    with pytest.raises(UnexpectedCellValueError):
        PIBExtractor.extract(strict)

    inspection = _inspection()
    activity = inspection.tables[0].activities[0]
    inspection.tables[0].activities = [activity, activity]
    inspection.sheets[0].detected_tables = inspection.tables
    _workbook(path, 10)
    with pytest.raises(DuplicateExtractionError):
        PIBExtractor.extract(_input(path, inspection))


def test_output_is_deterministic_and_metadata_is_separate(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _workbook(path, 10)
    first = PIBExtractor.extract(_input(path))
    second = PIBExtractor.extract(_input(path))

    assert first.records == second.records
    assert first.records[0].raw_value == 10
    assert first.metadata.started_at.tzinfo is not None
    assert first.metadata.started_at != second.metadata.started_at


def test_inspector_to_extractor_uses_real_inspection_result(tmp_path: Path) -> None:
    path = tmp_path / "pib.xlsx"
    _inspected_workbook(path)

    inspection = PIBWorkbookInspector.inspect(path)
    result = PIBExtractor.extract(ExtractionInput(path, inspection))

    assert inspection.is_valid
    assert len(inspection.tables) == 18
    assert result.records
    assert result.records[0].period_original == "2005-I"
    assert result.records[0].status is None
    assert result.metadata.tables_processed == 18